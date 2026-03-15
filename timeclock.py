#!/usr/bin/env python3
"""
=======================================================================
  RASPBERRY PI EMPLOYEE TIME CLOCK SYSTEM
  Hardware : Raspberry Pi 4B  |  I2C 20x4 LCD  |  3x4 Matrix Keypad
  Version  : 2.1
=======================================================================
"""

import os
import sys
import json
import time
import signal
import smtplib
import logging
import datetime
import threading
from pathlib import Path
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import RPi.GPIO as GPIO
from RPLCD.i2c import CharLCD
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

import config

# ───────────────────────────────────────────────────────────────────
#  LOGGING SETUP
# ───────────────────────────────────────────────────────────────────
Path(config.DATA_DIR).mkdir(parents=True, exist_ok=True)
Path(config.EXPORT_DIR).mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    filename='/home/timeclock/timeclock/timeclock.log',
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
log = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────
#  LCD INITIALISATION
# ───────────────────────────────────────────────────────────────────
try:
    lcd = CharLCD(
        i2c_expander='PCF8574',
        address=config.LCD_I2C_ADDRESS,
        port=config.LCD_I2C_PORT,
        cols=config.LCD_COLS,
        rows=config.LCD_ROWS,
        dotsize=8,
        charmap='A02',
        auto_linebreaks=False,
        backlight_enabled=True,
    )
    lcd.clear()
    log.info('LCD initialised at address 0x%02X', config.LCD_I2C_ADDRESS)
except Exception as exc:
    log.critical('LCD init failed: %s', exc)
    sys.exit(1)

COLS = config.LCD_COLS


def _center(text: str) -> str:
    text = str(text)[:COLS]
    pad_total = COLS - len(text)
    pad_left = pad_total // 2
    pad_right = pad_total - pad_left
    return (' ' * pad_left) + text + (' ' * pad_right)


def lcd_show(lines):
    lcd.clear()
    for row, text in enumerate((lines + [''] * 4)[:4]):
        lcd.cursor_pos = (row, 0)
        lcd.write_string(_center(text))


def fmt_12hr(dt: datetime.datetime) -> str:
    hour = int(dt.strftime('%I'))
    minute = dt.strftime('%M')
    ampm = dt.strftime('%p')
    return f'{hour}:{minute} {ampm}'


# ───────────────────────────────────────────────────────────────────
#  IDLE CLOCK DISPLAY THREAD
# ───────────────────────────────────────────────────────────────────
_idle = threading.Event()
_idle.set()
_lcd_lock = threading.Lock()


def _clock_thread():
    while True:
        if _idle.is_set():
            now = datetime.datetime.now()
            with _lcd_lock:
                lcd_show([
                    fmt_12hr(now),
                    now.strftime('%m/%d/%Y'),
                    'Enter ID',
                    'Then Press #',
                ])
        time.sleep(1)


threading.Thread(target=_clock_thread, daemon=True).start()
log.info('Clock display thread started.')

# ───────────────────────────────────────────────────────────────────
#  GPIO / KEYPAD SETUP
# ───────────────────────────────────────────────────────────────────
GPIO.setmode(GPIO.BCM)
GPIO.setwarnings(False)

ROW_PINS = config.KEYPAD_ROW_PINS
COL_PINS = config.KEYPAD_COL_PINS
KP_LAYOUT = config.KEYPAD_LAYOUT

for pin in ROW_PINS:
    GPIO.setup(pin, GPIO.OUT)
    GPIO.output(pin, GPIO.HIGH)

for pin in COL_PINS:
    GPIO.setup(pin, GPIO.IN, pull_up_down=GPIO.PUD_UP)


def _read_key():
    for r_idx, row_pin in enumerate(ROW_PINS):
        GPIO.output(row_pin, GPIO.LOW)
        for c_idx, col_pin in enumerate(COL_PINS):
            if GPIO.input(col_pin) == GPIO.LOW:
                key = KP_LAYOUT[r_idx][c_idx]
                GPIO.output(row_pin, GPIO.HIGH)

                timeout = time.time() + 2.0
                while GPIO.input(col_pin) == GPIO.LOW:
                    if time.time() > timeout:
                        break
                    time.sleep(0.02)
                time.sleep(0.05)
                return key
        GPIO.output(row_pin, GPIO.HIGH)
    return None


# ───────────────────────────────────────────────────────────────────
#  PAY PERIOD HELPERS
# ───────────────────────────────────────────────────────────────────
_ANCHOR = datetime.date.fromisoformat(config.PAY_PERIOD_ANCHOR)


def get_pay_period(d: datetime.date):
    delta = (d - _ANCHOR).days
    offset = (delta // 14) * 14
    start = _ANCHOR + datetime.timedelta(days=offset)
    end = start + datetime.timedelta(days=13)
    return start, end


def period_label(start: datetime.date) -> str:
    end = start + datetime.timedelta(days=13)
    return f"{start.strftime('%m/%d/%Y')} to {end.strftime('%m/%d/%Y')}"


# ───────────────────────────────────────────────────────────────────
#  JSON STORAGE
# ───────────────────────────────────────────────────────────────────
def _log_path(period_start: datetime.date) -> str:
    return os.path.join(config.DATA_DIR, f'log_{period_start.isoformat()}.json')


def _load(period_start: datetime.date):
    p = _log_path(period_start)
    if os.path.exists(p):
        with open(p, 'r') as f:
            return json.load(f)
    return {}


def _save(period_start: datetime.date, data):
    with open(_log_path(period_start), 'w') as f:
        json.dump(data, f, indent=2)


# ───────────────────────────────────────────────────────────────────
#  CLOCK IN / OUT
# ───────────────────────────────────────────────────────────────────
def clock_event(code: str):
    emp = config.EMPLOYEES[code]
    name = emp['name']
    emp_id = emp['id']

    now = datetime.datetime.now()
    today = now.date()
    time_str = now.strftime('%H:%M:%S')
    date_key = today.isoformat()

    ps, _ = get_pay_period(today)
    data = _load(ps)

    if emp_id not in data:
        data[emp_id] = {'name': name, 'id': emp_id, 'days': {}}

    day = data[emp_id]['days'].setdefault(date_key, [])

    if len(day) % 2 == 0:
        day.append({'in': time_str})
        action = 'IN'
    else:
        day[-1]['out'] = time_str
        action = 'OUT'

    _save(ps, data)
    log.info('Clock %s: %s (%s) at %s', action, name, emp_id, now.strftime('%Y-%m-%d %H:%M:%S'))
    return name, action


# ───────────────────────────────────────────────────────────────────
#  HOURS + EXCEL EXPORT
# ───────────────────────────────────────────────────────────────────
_FMT = '%H:%M:%S'
_THIN = Side(border_style='thin', color='000000')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _day_hours(punches):
    total = 0.0
    for p in punches:
        if 'in' in p and 'out' in p:
            t_in = datetime.datetime.strptime(p['in'], _FMT)
            t_out = datetime.datetime.strptime(p['out'], _FMT)
            secs = (t_out - t_in).total_seconds()
            if secs > 0:
                total += secs / 3600.0
    return round(total, 2)


def _fmt_time(t_str: str) -> str:
    try:
        dt = datetime.datetime.strptime(t_str, _FMT)
        hour = int(dt.strftime('%I'))
        mins = dt.strftime('%M')
        ampm = dt.strftime('%p')
        return f'{hour}:{mins} {ampm}'
    except Exception:
        return ''


def _cell(ws, row, col, value='', bold=False, align='left'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    return c


def build_excel(period_start: datetime.date) -> str:
    period_end = period_start + datetime.timedelta(days=13)
    p_label = period_label(period_start)
    data = _load(period_start)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for _, emp_info in config.EMPLOYEES.items():
        eid = emp_info['id']
        name = emp_info['name']
        edata = data.get(eid, {'name': name, 'id': eid, 'days': {}})

        ws = wb.create_sheet(title=name[:31])

        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 2
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 2
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 2
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 11
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 11
        ws.column_dimensions['M'].width = 2
        ws.column_dimensions['N'].width = 11
        ws.column_dimensions['O'].width = 11

        _cell(ws, 1, 1, name, bold=True)
        _cell(ws, 1, 3, 'Pay Period')
        _cell(ws, 1, 4, p_label)

        headers = {4: 'Clock In', 6: 'Clock Out', 8: 'Clock In', 10: 'Clock Out', 12: 'Total Hours'}
        for col, lbl in headers.items():
            _cell(ws, 3, col, lbl, align='center')

        def write_week(week_start: datetime.date, start_row: int):
            days_order = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            week_hours = 0.0

            for offset, day_name in enumerate(days_order):
                day_date = week_start + datetime.timedelta(days=offset)
                date_key = day_date.isoformat()
                punches = edata['days'].get(date_key, [])
                day_total = _day_hours(punches)
                week_hours += day_total

                row = start_row + offset
                _cell(ws, row, 1, day_name)
                _cell(ws, row, 2, f'{day_date.month}/{day_date.day}/{day_date.year}')

                col_pairs = [(4, 6), (8, 10)]
                for pair_idx, (col_in, col_out) in enumerate(col_pairs):
                    if pair_idx < len(punches):
                        p = punches[pair_idx]
                        _cell(ws, row, col_in, _fmt_time(p.get('in', '')), align='center')
                        _cell(ws, row, col_out, _fmt_time(p.get('out', '')), align='center')

                if day_total > 0:
                    _cell(ws, row, 12, (day_total if day_total != int(day_total) else int(day_total)), align='center')

            total_row = start_row + 7
            reg_hours = min(week_hours, 40.0)
            ot_hours = max(0.0, round(week_hours - 40.0, 2))

            _cell(ws, total_row, 11, 'Total', align='right')
            _cell(ws, total_row, 12, (reg_hours if reg_hours != int(reg_hours) else int(reg_hours)), align='center')
            _cell(ws, total_row, 14, 'Over Time', align='left')
            if ot_hours > 0:
                _cell(ws, total_row, 15, (ot_hours if ot_hours != int(ot_hours) else int(ot_hours)), align='center')

            return reg_hours, ot_hours

        week1_start = period_start
        week2_start = period_start + datetime.timedelta(days=7)
        w1_reg, w1_ot = write_week(week1_start, 4)
        w2_reg, w2_ot = write_week(week2_start, 13)

        grand_total = round(w1_reg + w2_reg, 2)
        grand_ot = round(w1_ot + w2_ot, 2)

        _cell(ws, 23, 11, 'Grand Totals', align='right')
        _cell(ws, 23, 12, (grand_total if grand_total != int(grand_total) else int(grand_total)), align='center')
        if grand_ot > 0:
            _cell(ws, 23, 15, (grand_ot if grand_ot != int(grand_ot) else int(grand_ot)), align='center')

    fname = f"TimeSheet_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.xlsx"
    fpath = os.path.join(config.EXPORT_DIR, fname)
    wb.save(fpath)
    log.info('Excel saved: %s', fpath)
    return fpath


# ───────────────────────────────────────────────────────────────────
#  EXPORT DESTINATIONS (SMB + Email)
# ────────────────────────────────  ──────────────────────────────────
def send_smb(file_path: str) -> bool:
    fname = os.path.basename(file_path)
    cmd = (
        f'smbclient \"//{config.SMB_SERVER}/{config.SMB_SHARE}\" '
        f'\"{config.SMB_PASSWORD}\" '
        f'-U \"{config.SMB_USERNAME}\" '
        f'-c \"put \\\\\"{file_path}\\\\\" \\\\\"{fname}\\\\\"\"'
    )
    rc = os.system(cmd)
    if rc == 0:
        log.info('SMB transfer OK -> //%s/%s/%s', config.SMB_SERVER, config.SMB_SHARE, fname)
        return True
    log.error('SMB transfer failed (rc=%d)', rc)
    return False


def send_email(file_path: str, period_start: datetime.date) -> bool:
    p_label = period_label(period_start)
    subject = f'Time Sheet Export - {p_label}'
    body = (
        f'Pay period time sheet attached.\n\n'
        f'Period : {p_label}\n'
        f'Created: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n'
        f'(Automated message from Employee Time Clock System)'
    )

    msg = MIMEMultipart()
    msg['From'] = config.EMAIL_SENDER
    msg['To'] = ', '.join(config.EMAIL_RECIPIENTS)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    with open(file_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
            s.login(config.EMAIL_SENDER, config.EMAIL_APP_PASS)
            s.sendmail(config.EMAIL_SENDER, config.EMAIL_RECIPIENTS, msg.as_string())
        log.info('Email sent to: %s', config.EMAIL_RECIPIENTS)
        return True
    except Exception as exc:
        log.error('Email failed: %s', exc)
        return False


def full_export(period_start: datetime.date):
    with _lcd_lock:
        lcd_show(['', 'Exporting...', 'Please wait', ''])
    try:
        fpath = build_excel(period_start)
        smb_ok = send_smb(fpath)
        email_ok = send_email(fpath, period_start)

        if smb_ok and email_ok:
            result = ['Export Complete!', 'File sent to PC', '& emailed!', '']
        elif smb_ok:
            result = ['Export Partial', 'Sent to PC OK', 'Email FAILED', 'Check logs']
        elif email_ok:
            result = ['Export Partial', 'Email sent OK', 'PC xfer FAILED', 'Check logs']
        else:
            result = ['EXPORT FAILED', 'Check network', 'and email cfg', 'See logs']

        with _lcd_lock:
            lcd_show(result)
        time.sleep(4)
    except Exception as exc:
        log.error('full_export error: %s', exc, exc_info=True)
        with _lcd_lock:
            lcd_show(['EXPORT ERROR', str(exc)[:20], 'Check logs', ''])
        time.sleep(4)


# ───────────────────────────────────────────────────────────────────
#  AUTO EXPORT
# ───────────────────────────────────────────────────────────────────
_last_auto_export = None


def _check_auto_export():
    global _last_auto_export
    now = datetime.datetime.now()
    today = now.date()

    # Friday 23:59
    if today.weekday() == 4 and now.hour == 23 and now.minute == 59 and _last_auto_export != today:
        _last_auto_export = today
        ps, _ = get_pay_period(today)
        log.info('Auto-export triggered for period starting %s', ps)
        threading.Thread(target=full_export, args=(ps,), daemon=True).start()


# ───────────────────────────────────────────────────────────────────
#  KEYPAD INPUT UI
# ───────────────────────────────────────────────────────────────────
def get_employee_code() -> str:
    digits = []
    first_key_pressed = False

    def refresh():
        masked = ('*' * len(digits)).ljust(4, '_')
        with _lcd_lock:
            lcd_show(['Enter ID:', masked, '[*]=Clear', '[#]=Submit'])

    while True:
        key = _read_key()
        if key is None:
            time.sleep(0.05)
            continue

        if not first_key_pressed:
            first_key_pressed = True
            _idle.clear()
            refresh()

        if key == config.KEY_CLEAR:
            digits.clear()
            refresh()

        elif key == config.KEY_ENTER:
            if len(digits) == 4:
                return ''.join(digits)
            with _lcd_lock:
                lcd_show(['Need 4 digits', f'You entered: {len(digits)}', 'Press * to clear', 'then try again'])
            time.sleep(2)
            digits.clear()
            refresh()

        elif key.isdigit():
            if len(digits) < 4:
                digits.append(key)
                refresh()

        time.sleep(0.05)


# ───────────────────────────────────────────────────────────────────
#  GRACEFUL SHUTDOWN
# ───────────────────────────────────────────────────────────────────
def _shutdown(signum, frame):
    log.info('Shutdown signal received.')
    with _lcd_lock:
        lcd_show(['', 'Shutting Down...', '', ''])
    time.sleep(1)
    lcd.close(clear=True)
    GPIO.cleanup()
    sys.exit(0)


signal.signal(signal.SIGTERM, _shutdown)
signal.signal(signal.SIGINT, _shutdown)


# ───────────────────────────────────────────────────────────────────
#  MAIN LOOP
# ───────────────────────────────────────────────────────────────────
def main():
    log.info('=' * 60)
    log.info('Employee Time Clock System STARTED v2.1')
    log.info('=' * 60)

    while True:
        try:
            _check_auto_export()

            _idle.set()
            time.sleep(0.5)

            code = get_employee_code()

            # Admin export
            if code == config.EXPORT_CODE:
                ps, _ = get_pay_period(datetime.date.today())
                with _lcd_lock:
                    lcd_show(['ADMIN EXPORT', 'Current Period', 'Starting...', ''])
                time.sleep(1)
                full_export(ps)
                continue

            # Employee
            if code in config.EMPLOYEES:
                emp_name, action = clock_event(code)
                action_lbl = 'CLOCKED  IN' if action == 'IN' else 'CLOCKED OUT'
                now = datetime.datetime.now()
                with _lcd_lock:
                    lcd_show([emp_name[:20], action_lbl, fmt_12hr(now), ''])
                time.sleep(4)
            else:
                with _lcd_lock:
                    lcd_show(['INVALID CODE', 'Please try again', '', ''])
                log.warning('Invalid code entered: %s', code)
                time.sleep(3)

        except Exception as exc:
            log.error('Main loop error: %s', exc, exc_info=True)
            with _lcd_lock:
                lcd_show(['SYSTEM ERROR', str(exc)[:20], 'Restarting...', ''])
            time.sleep(5)
        finally:
            _idle.set()


if __name__ == '__main__':
    main()